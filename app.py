from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session, send_from_directory
import os
from pdf2image import convert_from_path
from PIL import Image, ImageChops
import shutil
from werkzeug.utils import secure_filename
import pandas as pd
import xlsxwriter
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import atexit
from threading import Thread
from time import sleep
import requests
import subprocess
import tempfile
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

# Initialize Flask app
app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = 'your-secret-key-123'  # Change this for production

# Configuration
app.config.update(
    UPLOAD_FOLDER_PDF='static/uploads/pdfs',
    OUTPUT_FOLDER_PNG='static/outputs/pngs',
    UPLOAD_FOLDER_EXCEL='static/uploads/excels',
    OUTPUT_FOLDER_REPORTS='static/outputs/reports',
    UPLOAD_FOLDER_SALES_DATA='static/uploads/sales_data',
    OUTPUT_FOLDER_SALES_REPORTS='static/outputs/sales_reports',
    UPLOAD_FOLDER_VBA='static/uploads/vba_pdfs',
    VBA_OUTPUT_FOLDER='static/outputs/vba_output',
    VBA_TEMPLATE_PATH='static/templates/VBAPdfExportTemplate.xlsm',
    ALLOWED_EXTENSIONS_PDF={'pdf'},
    ALLOWED_EXTENSIONS_EXCEL={'xlsx', 'xls', 'xlsm'},
    MAX_CONTENT_LENGTH=16 * 1024 * 1024  # 16MB max upload
)

def clear_all_folders():
    """Clear all upload and output folders"""
    folders_to_clear = [
        app.config['UPLOAD_FOLDER_PDF'],
        app.config['OUTPUT_FOLDER_PNG'],
        app.config['UPLOAD_FOLDER_EXCEL'],
        app.config['OUTPUT_FOLDER_REPORTS'],
        app.config['UPLOAD_FOLDER_SALES_DATA'],
        app.config['OUTPUT_FOLDER_SALES_REPORTS'],
        app.config['UPLOAD_FOLDER_VBA'],
        app.config['VBA_OUTPUT_FOLDER']
    ]

    for folder in folders_to_clear:
        try:
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f'Failed to delete {file_path}. Reason: {e}')
        except FileNotFoundError:
            os.makedirs(folder, exist_ok=True)

# Clear folders on application start
clear_all_folders()

# Register cleanup function to run when application exits
atexit.register(clear_all_folders)

# Create folders if they don't exist
for folder in [
    app.config['UPLOAD_FOLDER_PDF'],
    app.config['OUTPUT_FOLDER_PNG'],
    app.config['UPLOAD_FOLDER_EXCEL'],
    app.config['OUTPUT_FOLDER_REPORTS'],
    app.config['UPLOAD_FOLDER_SALES_DATA'],
    app.config['OUTPUT_FOLDER_SALES_REPORTS'],
    app.config['UPLOAD_FOLDER_VBA'],
    app.config['VBA_OUTPUT_FOLDER'],
    os.path.dirname(app.config['VBA_TEMPLATE_PATH'])
]:
    os.makedirs(folder, exist_ok=True)
# Clear folders on application start
clear_all_folders()

# Register cleanup function to run when application exits
atexit.register(clear_all_folders)

# Create folders if they don't exist
for folder in [
    app.config['UPLOAD_FOLDER_PDF'],
    app.config['OUTPUT_FOLDER_PNG'],
    app.config['UPLOAD_FOLDER_EXCEL'],
    app.config['OUTPUT_FOLDER_REPORTS'],
    app.config['UPLOAD_FOLDER_SALES_DATA'],
    app.config['OUTPUT_FOLDER_SALES_REPORTS'],
    app.config['UPLOAD_FOLDER_VBA'],
    os.path.dirname(app.config['VBA_TEMPLATE_PATH'])
]:
    os.makedirs(folder, exist_ok=True)

def allowed_file(filename, file_type='pdf'):
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    if file_type == 'pdf':
        return ext in app.config['ALLOWED_EXTENSIONS_PDF']
    return ext in app.config['ALLOWED_EXTENSIONS_EXCEL']

def trim_whitespace(image):
    bg = Image.new(image.mode, image.size, (255, 255, 255))
    diff = ImageChops.difference(image, bg)
    bbox = diff.getbbox()
    return image.crop(bbox) if bbox else image

def has_png_files():
    return any(fname.endswith('.png') for fname in os.listdir(app.config['OUTPUT_FOLDER_PNG']))

def generate_png_name(base_name, output_folder):
    """Generate unique PNG filename with incremental suffix if needed"""
    counter = 1
    name = f"{base_name}.png"
    while os.path.exists(os.path.join(output_folder, name)):
        name = f"{base_name}_{counter}.png"
        counter += 1
    return name

def safe_divide(numerator, denominator):
    """Safe division that handles division by zero"""
    if denominator == 0:
        return float('inf') if numerator > 0 else float('-inf') if numerator < 0 else 0
    return numerator / denominator

def format_change(current, previous, is_amount=False):
    try:
        if is_amount:
            current = round(current)
            previous = round(previous)
            cur_str = f"{current:,}"
            prev_str = f"{previous:,}"
        else:
            cur_str = f"{current}"
            prev_str = f"{previous}"

        if current == previous:
            return f"{cur_str} (No change)"

        if previous == 0:
            return f"{cur_str} (New data)"

        percent = round(safe_divide((current - previous), abs(previous)) * 100)

        if current > previous:
            return f"{cur_str} (↑ {percent}% from {prev_str})"
        else:
            return f"{cur_str} (↓ {abs(percent)}% from {prev_str})"
    except Exception:
        return f"{current} (Error calculating change)"

def validate_excel_data(df):
    """Validate the Excel data structure and content"""
    required_columns = [
        'Sales Person',
        'Current Bid participate',
        'Last Bid participate',
        'Current Bid Visitor',
        'Last Bid Visitor',
        'Current Bid Winner',
        'Last Bid Winner',
        'Curr Bid Amt(Cnf)',
        'Last Bid Amt(Cnf)',
        'Current Bid Amt(All)',
        'Last Bid Amt(All)'
    ]

    # Check if all required columns exist
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Convert numeric columns to appropriate types
    numeric_columns = [
        'Current Bid participate', 'Last Bid participate',
        'Current Bid Visitor', 'Last Bid Visitor',
        'Current Bid Winner', 'Last Bid Winner',
        'Curr Bid Amt(Cnf)', 'Last Bid Amt(Cnf)',
        'Current Bid Amt(All)', 'Last Bid Amt(All)'
    ]

    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    return df

def generate_excel_report(input_path):
    try:
        # Read and validate the Excel file
        df = pd.read_excel(input_path)
        df = validate_excel_data(df)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(app.config['OUTPUT_FOLDER_REPORTS'], f"Bids_summary_{timestamp}.xlsx")

        workbook = xlsxwriter.Workbook(output_file)
        worksheet = workbook.add_worksheet("Summary")

        # Formats
        bold = workbook.add_format({'bold': True})
        normal = workbook.add_format()
        wrap = workbook.add_format({'text_wrap': True})
        error_format = workbook.add_format({'font_color': 'red'})

        row = 0
        for _, r in df.iterrows():
            try:
                name = str(r['Sales Person'])
                name_line = name if name.strip().lower().startswith("mr.") else f"Mr. {name}"

                summary_lines = [
                    [bold, f"{name_line}\n"],
                    [bold, "• Bid Participation: "], normal,
                    format_change(r['Current Bid participate'], r['Last Bid participate']) + "\n",
                    [bold, "• Bid Visitors: "], normal,
                    format_change(r['Current Bid Visitor'], r['Last Bid Visitor']) + "\n",
                    [bold, "• Bid Wins: "], normal, format_change(r['Current Bid Winner'], r['Last Bid Winner']) + "\n",
                    [bold, "• Confirmed Bid Amount: "], normal,
                    format_change(r['Curr Bid Amt(Cnf)'], r['Last Bid Amt(Cnf)'], is_amount=True) + "\n",
                    [bold, "• Total Bid Amount (All): "], normal,
                    format_change(r['Current Bid Amt(All)'], r['Last Bid Amt(All)'], is_amount=True),
                ]

                rich_text = []
                for i in summary_lines:
                    if isinstance(i, list):
                        rich_text.append(i[0])
                        rich_text.append(i[1])
                    else:
                        rich_text.append(i)

                worksheet.write_rich_string(row, 0, *rich_text, wrap)
                row += 1
            except Exception as e:
                worksheet.write_string(row, 0, f"Error processing row {row + 1}: {str(e)}", error_format)
                row += 1

        worksheet.set_column(0, 0, 90)
        workbook.close()
        return output_file
    except Exception as e:
        raise Exception(f"Error generating report: {str(e)}")

def generate_sales_performance_report(input_path):
    try:
        # Read Excel file
        df = pd.read_excel(input_path, sheet_name="Sheet2")

        # Set the first column as index (metrics)
        df.set_index(df.columns[0], inplace=True)

        # Clean column names
        df.columns = [col.strip() for col in df.columns]

        # Function to format values
        def format_value(value, metric):
            if '%' in metric or 'Ratio' in metric:
                return f"{float(value):.2%}"
            elif metric in ['New Gain Customers', 'Total Business', 'Adding Bid', 'Avg Days of stone sold ']:
                return str(int(float(value)))
            elif isinstance(value, float):
                return f"{value:.2f}"
            return str(value)

        # Function to generate clean reports
        def generate_report(person_name):
            relevant_metrics = [
                'New Gain Customers',
                'Total Business',
                'Adding Bid',
                'Customer Active Ratio',
                'Sale Amount % Through Bid ',
                '% of Sale from top 10 customer',
                'Avg Days of stone sold ',
                'Goal Achvd %'
            ]

            report_lines = [
                f"Performance Summary - {person_name}",
                "-" * 40
            ]

            for metric in relevant_metrics:
                if metric in df.index:
                    value = df.loc[metric, person_name]
                    formatted_value = format_value(value, metric)
                    report_lines.append(f"{metric}: {formatted_value}")

            return "\n".join(report_lines)

        # Generate all reports
        all_reports = {person: generate_report(person) for person in df.columns[1:]}

        # Create output filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(app.config['OUTPUT_FOLDER_SALES_REPORTS'],
                                 f"Sales_Performance_Reports_{timestamp}.xlsx")

        # Create a DataFrame for Excel output
        output_data = []
        for person, report in all_reports.items():
            output_data.append([person, report])

        output_df = pd.DataFrame(output_data, columns=['Salesperson', 'Performance Report'])

        # Export to single Excel sheet with perfect formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            output_df.to_excel(writer, sheet_name='Performance Reports', index=False)

            # Get the worksheet and apply formatting
            worksheet = writer.sheets['Performance Reports']

            # Set column widths
            worksheet.column_dimensions['A'].width = 20  # Salesperson column
            worksheet.column_dimensions['B'].width = 60  # Wider report column

            # Apply formatting to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    cell.font = Font(name='Calibri', size=11)

            # Freeze header row
            worksheet.freeze_panes = "A2"

        return output_file
    except Exception as e:
        raise Exception(f"Error generating sales performance report: {str(e)}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    # Check for PDF files
    pdf_files = []
    pdf_folder = app.config['UPLOAD_FOLDER_PDF']
    if os.path.exists(pdf_folder):
        pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')]

    # Check for Excel files
    excel_files = []
    excel_folder = app.config['UPLOAD_FOLDER_EXCEL']
    if os.path.exists(excel_folder):
        excel_files = [f for f in os.listdir(excel_folder) if f.lower().endswith(('.xlsx', '.xls'))]

    # Check for report files
    report_files = []
    report_folder = app.config['OUTPUT_FOLDER_REPORTS']
    if os.path.exists(report_folder):
        report_files = [f for f in os.listdir(report_folder) if f.lower().endswith('.xlsx')]

    # Check for VBA PDFs
    vba_pdfs = []
    vba_folder = app.config['UPLOAD_FOLDER_VBA']
    if os.path.exists(vba_folder):
        vba_pdfs = [f for f in os.listdir(vba_folder) if f.lower().endswith('.pdf')]
        
    # Check for VBA generated PDFs
    vba_generated_pdfs = []
    vba_output_folder = os.path.join(app.config['VBA_OUTPUT_FOLDER'], 'temp')
    if os.path.exists(vba_output_folder):
        vba_generated_pdfs = [f for f in os.listdir(vba_output_folder) if f.lower().endswith('.pdf')]

    # Check for sales data files
    sales_data_files = []
    sales_data_folder = app.config['UPLOAD_FOLDER_SALES_DATA']
    if os.path.exists(sales_data_folder):
        sales_data_files = [f for f in os.listdir(sales_data_folder) if f.lower().endswith(('.xlsx', '.xls'))]

    # Check for sales report files
    sales_report_files = []
    sales_report_folder = app.config['OUTPUT_FOLDER_SALES_REPORTS']
    if os.path.exists(sales_report_folder):
        sales_report_files = [f for f in os.listdir(sales_report_folder) if f.lower().endswith('.xlsx')]

    # Check if any PNGs exist
    has_pngs = False
    png_folder = app.config['OUTPUT_FOLDER_PNG']
    if os.path.exists(png_folder):
        has_pngs = any(f.lower().endswith('.png') for f in os.listdir(png_folder))

    return render_template('dashboard.html',
                         pdf_files=pdf_files,
                         excel_files=excel_files,
                         report_files=report_files,
                         sales_data_files=sales_data_files,
                         sales_report_files=sales_report_files,
                         vba_pdfs=vba_pdfs,
                         vba_generated_pdfs=vba_generated_pdfs,  # Add this line
                         has_pngs=has_pngs)

@app.route('/upload-pdf', methods=['POST'])
def upload_pdf():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('dashboard'))

    # Clear upload folder first
    for filename in os.listdir(app.config['UPLOAD_FOLDER_PDF']):
        file_path = os.path.join(app.config['UPLOAD_FOLDER_PDF'], filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            flash(f'Error clearing upload folder: {e}', 'error')

    files = request.files.getlist('file')
    file_count = 0

    for file in files:
        if file.filename == '':
            continue

        if file and allowed_file(file.filename, 'pdf'):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER_PDF'], filename))
            file_count += 1

    if file_count > 0:
        flash(f'Successfully uploaded {file_count} PDF file(s)', 'success')
    else:
        flash('No valid PDF files uploaded', 'error')

    return redirect(url_for('dashboard'))

@app.route('/convert-pdf', methods=['POST'])
def convert_pdf():
    # Clear output folder first
    for filename in os.listdir(app.config['OUTPUT_FOLDER_PNG']):
        file_path = os.path.join(app.config['OUTPUT_FOLDER_PNG'], filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            flash(f'Error clearing output folder: {e}', 'error')

    pdf_files = os.listdir(app.config['UPLOAD_FOLDER_PDF'])
    success_count = 0

    for filename in pdf_files:
        if filename.endswith('.pdf'):
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER_PDF'], filename)
            try:
                images = convert_from_path(pdf_path, dpi=300)
                base_name = os.path.splitext(filename)[0]

                for img in images:
                    trimmed = trim_whitespace(img)
                    png_name = generate_png_name(base_name, app.config['OUTPUT_FOLDER_PNG'])
                    trimmed.save(os.path.join(app.config['OUTPUT_FOLDER_PNG'], png_name))

                success_count += 1
                os.remove(pdf_path)
            except Exception as e:
                flash(f'Error converting {filename}: {e}', 'error')

    if success_count > 0:
        flash(f'Successfully converted {success_count} PDF file(s) to PNGs', 'success')

    return redirect(url_for('dashboard'))

@app.route('/download-pngs')
def download_pngs():
    zip_filename = 'converted_images.zip'
    try:
        shutil.make_archive("converted_images", 'zip', app.config['OUTPUT_FOLDER_PNG'])
        return send_file(zip_filename, as_attachment=True)
    except Exception as e:
        flash(f'Error creating ZIP file: {e}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('dashboard'))

    # Clear upload folder first
    for filename in os.listdir(app.config['UPLOAD_FOLDER_EXCEL']):
        file_path = os.path.join(app.config['UPLOAD_FOLDER_EXCEL'], filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            flash(f'Error clearing upload folder: {e}', 'error')

    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('dashboard'))

    if file and allowed_file(file.filename, 'excel'):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER_EXCEL'], filename))
        flash('Excel file uploaded successfully!', 'success')
    else:
        flash('Invalid file type. Only Excel files are allowed', 'error')

    return redirect(url_for('dashboard'))

@app.route('/generate-report', methods=['POST'])
def generate_report():
    if not os.listdir(app.config['UPLOAD_FOLDER_EXCEL']):
        flash('No Excel files uploaded to generate report', 'error')
        return redirect(url_for('dashboard'))

    try:
        excel_file = os.listdir(app.config['UPLOAD_FOLDER_EXCEL'])[0]
        input_path = os.path.join(app.config['UPLOAD_FOLDER_EXCEL'], excel_file)
        output_file = generate_excel_report(input_path)
        flash('Report generated successfully!', 'success')
        session['latest_report'] = output_file

        # Remove the Excel file after successful report generation
        os.remove(input_path)
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')

    return redirect(url_for('dashboard'))

@app.route('/download-report')
def download_report():
    if 'latest_report' not in session or not os.path.exists(session['latest_report']):
        flash('No report available for download', 'error')
        return redirect(url_for('dashboard'))

    return send_file(session['latest_report'], as_attachment=True)

@app.route('/upload-sales-data', methods=['POST'])
def upload_sales_data():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('dashboard'))

    # Clear upload folder first
    for filename in os.listdir(app.config['UPLOAD_FOLDER_SALES_DATA']):
        file_path = os.path.join(app.config['UPLOAD_FOLDER_SALES_DATA'], filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            flash(f'Error clearing upload folder: {e}', 'error')

    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('dashboard'))

    if file and allowed_file(file.filename, 'excel'):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER_SALES_DATA'], filename))
        flash('Sales data file uploaded successfully!', 'success')
    else:
        flash('Invalid file type. Only Excel files are allowed', 'error')

    return redirect(url_for('dashboard'))

@app.route('/generate-sales-report', methods=['POST'])
def generate_sales_report():
    if not os.listdir(app.config['UPLOAD_FOLDER_SALES_DATA']):
        flash('No sales data files uploaded to generate report', 'error')
        return redirect(url_for('dashboard'))

    try:
        sales_data_file = os.listdir(app.config['UPLOAD_FOLDER_SALES_DATA'])[0]
        input_path = os.path.join(app.config['UPLOAD_FOLDER_SALES_DATA'], sales_data_file)
        output_file = generate_sales_performance_report(input_path)
        flash('Sales performance report generated successfully!', 'success')
        session['latest_sales_report'] = output_file

        # Remove the sales data file after successful report generation
        os.remove(input_path)
    except Exception as e:
        flash(f'Error generating sales performance report: {str(e)}', 'error')

    return redirect(url_for('dashboard'))

@app.route('/download-sales-report')
def download_sales_report():
    if 'latest_sales_report' not in session or not os.path.exists(session['latest_sales_report']):
        flash('No sales report available for download', 'error')
        return redirect(url_for('dashboard'))

    return send_file(session['latest_sales_report'], as_attachment=True)

# Replace the VBA Excel processing with this cross-platform version
@app.route('/process-vba-excel', methods=['POST'])
def process_vba_excel():
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('dashboard'))

    if file and allowed_file(file.filename, 'excel'):
        try:
            # Clear previous outputs
            temp_output = os.path.join(app.config['VBA_OUTPUT_FOLDER'], 'temp')
            if os.path.exists(temp_output):
                shutil.rmtree(temp_output)
            os.makedirs(temp_output, exist_ok=True)

            # Save the uploaded file
            filename = secure_filename(file.filename)
            input_path = os.path.join(app.config['UPLOAD_FOLDER_VBA'], filename)
            file.save(input_path)
            
            # Process the Excel file using openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(input_path)
            ws = wb.active  # Get the first worksheet

            generated_pdfs = []
            
            # Method 1: Export each column as separate PDF (like first VBA macro)
            if ws.max_column >= 3:  # At least column C exists
                for col in range(3, min(23, ws.max_column + 1)):  # Columns C to V (3 to 22)
                    # Get column header for filename
                    cell_value = str(ws.cell(row=1, column=col).value).strip()
                    if not cell_value:
                        cell_value = f"Col_{col}"
                    
                    # Create PDF
                    pdf_name = f"{cell_value}.pdf"
                    pdf_path = os.path.join(temp_output, pdf_name)
                    
                    # Create PDF using reportlab with table
                    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
                    data = []
                    
                    # Add header
                    data.append([cell_value])
                    
                    # Add column data (first 20 rows)
                    for row in range(2, min(22, ws.max_row + 1)):
                        cell_value = str(ws.cell(row=row, column=col).value)
                        if cell_value:
                            data.append([cell_value])
                    
                    # Create table
                    table = Table(data, colWidths=[500], rowHeights=20)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    doc.build([table])
                    generated_pdfs.append(pdf_name)
            
            # Method 2: Export specific ranges (like second VBA macro for bid summary)
            if ws.title == "Screen Shot":  # Check if this is the bid summary sheet
                range_definitions = [
                    ("A1:BM2", "Summary_Header"),
                    ("A10:BM11", ws['B10'].value),
                    ("A13:BM14", ws['B13'].value),
                    ("A16:BM17", ws['B16'].value),
                    ("A19:BM20", ws['B19'].value),
                    ("A22:BM23", ws['B22'].value),
                    ("A25:BM26", ws['B25'].value),
                    ("A28:BM29", ws['B28'].value),
                    ("A31:BM32", ws['B31'].value),
                    ("A34:BM35", ws['B34'].value),
                    ("A37:BM38", ws['B37'].value),
                    ("A4:BM5", "Summary_Footer"),
                    ("A40:BM41", ws['B40'].value),
                    ("A43:BM44", ws['B43'].value),
                    ("A46:BM47", ws['B46'].value),
                    ("A49:BM50", ws['B49'].value),
                    ("A52:BM53", ws['B52'].value),
                    ("A55:BM56", ws['B55'].value),
                    ("A7:BM8", "Summary_Details")
                ]
                
                for rng, name in range_definitions:
                    if not name:
                        continue
                        
                    pdf_name = f"{name}.pdf"
                    pdf_path = os.path.join(temp_output, pdf_name)
                    
                    # Parse the range
                    start_col, start_row, end_col, end_row = parse_range(rng)
                    
                    # Extract data from the range
                    data = []
                    for row in range(start_row, end_row + 1):
                        row_data = []
                        for col in range(start_col, end_col + 1):
                            cell = ws.cell(row=row, column=col)
                            row_data.append(str(cell.value) if cell.value else "")
                        data.append(row_data)
                    
                    # Create PDF with table
                    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
                    table = Table(data)
                    
                    # Add style
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('WORDWRAP', (0, 0), (-1, -1)),
                    ]))
                    
                    doc.build([table])
                    generated_pdfs.append(pdf_name)
            
            if not generated_pdfs:
                flash('No PDFs were generated', 'warning')
            else:
                session['vba_generated_pdfs'] = [os.path.basename(p) for p in generated_pdfs]
                flash(f'Successfully generated {len(generated_pdfs)} PDF file(s)', 'success')
            
            # Clean up
            wb.close()
            os.remove(input_path)
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            app.logger.error(f"Error in process_vba_excel: {str(e)}")
    else:
        flash('Invalid file type. Only Excel files allowed', 'error')

    return redirect(url_for('dashboard'))

def parse_range(rng_str):
    """Parse Excel range string into column/row numbers"""
    import re
    # Split range like "A1:BM2" into ["A1", "BM2"]
    cells = rng_str.split(':')
    if len(cells) != 2:
        raise ValueError(f"Invalid range format: {rng_str}")
    
    # Parse first cell
    col1 = re.sub(r'\d', '', cells[0]).upper()
    row1 = int(re.sub(r'[A-Z]', '', cells[0]))
    
    # Parse second cell
    col2 = re.sub(r'\d', '', cells[1]).upper()
    row2 = int(re.sub(r'[A-Z]', '', cells[1]))
    
    # Convert column letters to numbers (A=1, B=2, ..., Z=26, AA=27, etc.)
    def col_to_num(col):
        num = 0
        for c in col:
            num = num * 26 + (ord(c) - ord('A') + 1)
        return num
    
    return (col_to_num(col1), row1, col_to_num(col2), row2)

@app.route('/download-vba-pdf/<filename>')
def download_vba_pdf(filename):
    try:
        return send_from_directory(
            os.path.join(app.config['VBA_OUTPUT_FOLDER'], 'temp'),
            filename,
            as_attachment=True
        )
    except FileNotFoundError:
        flash('Requested PDF file not found', 'error')
        return redirect(url_for('dashboard'))

@app.route('/download-all-vba-pdfs')
def download_all_vba_pdfs():
    temp_output = os.path.join(app.config['VBA_OUTPUT_FOLDER'], 'temp')
    if not os.path.exists(temp_output):
        flash('No PDFs available for download', 'error')
        return redirect(url_for('dashboard'))
    
    zip_filename = 'vba_generated_pdfs.zip'
    zip_path = os.path.join(app.config['VBA_OUTPUT_FOLDER'], zip_filename)
    
    try:
        shutil.make_archive(
            os.path.join(app.config['VBA_OUTPUT_FOLDER'], 'vba_generated_pdfs'), 
            'zip', 
            temp_output
        )
        return send_file(zip_path, as_attachment=True)
    except Exception as e:
        flash(f'Error creating ZIP file: {e}', 'error')
        return redirect(url_for('dashboard'))

def ping_self():
    while True:
        try:
            # Ping your own Render URL every 5 minutes
            requests.get("https://your-app-name.onrender.com")
            sleep(300)  # 300 seconds = 5 minutes
        except:
            sleep(60)  # If error, wait 1 minute and retry

# Start keep-alive thread when not in debug mode
if not app.debug and os.environ.get("WERKZEUG_RUN_MAIN") != "true":
    t = Thread(target=ping_self)
    t.daemon = True
    t.start()

if __name__ == '__main__':
    app.run(debug=True)
